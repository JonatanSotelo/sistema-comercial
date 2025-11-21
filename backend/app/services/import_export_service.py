# app/services/import_export_service.py
from typing import Dict, List, Any, Optional, Tuple
from io import BytesIO, StringIO
import csv
import re

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _digits(s: str | None) -> str | None:
    """Normaliza string a solo dígitos (para CUIT/teléfono)"""
    if not s:
        return None
    return re.sub(r"\D+", "", str(s)) or None


def _normalize_decimal(value: Any) -> float | None:
    """Normaliza decimal (acepta coma o punto, elimina separadores)"""
    if value is None:
        return None
    s = str(value).strip()
    # Reemplazar coma por punto si hay ambos, o solo coma
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return None


def _normalize_int(value: Any) -> int | None:
    """Normaliza entero (elimina puntos/comas)"""
    if value is None:
        return None
    s = str(value).strip().replace(".", "").replace(",", "")
    try:
        return int(float(s))
    except:
        return None


def _map_column_name(col: str) -> str:
    """Mapea alias de columnas comunes"""
    col_lower = col.lower().strip()
    mapping = {
        "name": "nombre",
        "telefono": "telefono",
        "phone": "telefono",
        "phone_number": "telefono",
        "cuit": "cuit",
        "tax_id": "cuit",
        "dni": "cuit",
        "documento": "cuit",
        "id_number": "cuit",
        "national_id": "cuit",
        "cuil": "cuit",
        "direccion": "direccion",
        "address": "direccion",
        "email": "email",
        "correo": "email",
        "precio": "precio",
        "price": "precio",
        "stock": "stock",
        "proveedor_id": "proveedor_id",
        "proveedor": "proveedor_id",
        "supplier_id": "proveedor_id",
    }
    return mapping.get(col_lower, col_lower)


def parse_csv_file(file_content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Parsea archivo CSV y devuelve (headers, rows)"""
    text = file_content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    headers = [h.strip() for h in reader.fieldnames or []]
    rows = []
    for row in reader:
        rows.append({k.strip(): v for k, v in row.items() if k.strip()})
    return headers, rows


def parse_xlsx_file(file_content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Parsea archivo XLSX y devuelve (headers, rows)"""
    if not HAS_OPENPYXL:
        raise ValueError("openpyxl no está instalado")
    wb = load_workbook(BytesIO(file_content), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_dict = {}
        for i, cell in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = cell.value
        if any(row_dict.values()):
            rows.append(row_dict)
    return headers, rows


def parse_import_file(file_content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Parsea archivo CSV o XLSX según extensión"""
    if filename.endswith(".csv"):
        return parse_csv_file(file_content)
    elif filename.endswith((".xlsx", ".xls")):
        return parse_xlsx_file(file_content)
    else:
        raise ValueError(f"Formato no soportado: {filename}")


def normalize_cliente_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Normaliza fila de cliente con alias y validaciones"""
    normalized = {}
    
    # Nombre (requerido)
    nombre = row.get("nombre") or row.get("name") or ""
    if nombre:
        normalized["nombre"] = str(nombre).strip()
    
    # Email (opcional, validar si viene)
    email = row.get("email") or row.get("correo") or ""
    if email:
        email = str(email).strip()
        if "@" in email:
            normalized["email"] = email
    
    # Teléfono (normalizar a dígitos)
    telefono = row.get("telefono") or row.get("phone") or row.get("phone_number") or ""
    if telefono:
        tel_ok = _digits(str(telefono))
        if tel_ok:
            normalized["telefono"] = tel_ok
    
    # CUIT (normalizar a dígitos)
    cuit = row.get("cuit") or row.get("tax_id") or row.get("dni") or row.get("documento") or row.get("id_number") or row.get("national_id") or row.get("cuil") or ""
    if cuit:
        cuit_ok = _digits(str(cuit))
        if cuit_ok:
            normalized["cuit"] = cuit_ok
    
    return normalized


def normalize_proveedor_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Normaliza fila de proveedor con alias y validaciones"""
    normalized = {}
    
    # Nombre (requerido)
    nombre = row.get("nombre") or row.get("name") or ""
    if nombre:
        normalized["nombre"] = str(nombre).strip()
    
    # Email (opcional, validar si viene)
    email = row.get("email") or row.get("correo") or ""
    if email:
        email = str(email).strip()
        if "@" in email:
            normalized["email"] = email
    
    # Teléfono (normalizar a dígitos)
    telefono = row.get("telefono") or row.get("phone") or row.get("phone_number") or ""
    if telefono:
        tel_ok = _digits(str(telefono))
        if tel_ok:
            normalized["telefono"] = tel_ok
    
    # CUIT (normalizar a dígitos)
    cuit = row.get("cuit") or row.get("tax_id") or row.get("dni") or row.get("documento") or row.get("id_number") or row.get("national_id") or row.get("cuil") or ""
    if cuit:
        cuit_ok = _digits(str(cuit))
        if cuit_ok:
            normalized["cuit"] = cuit_ok
    
    # Dirección (opcional)
    direccion = row.get("direccion") or row.get("address") or ""
    if direccion:
        normalized["direccion"] = str(direccion).strip()
    
    return normalized


def normalize_producto_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Normaliza fila de producto con alias y validaciones"""
    normalized = {}
    
    # Nombre (requerido)
    nombre = row.get("nombre") or row.get("name") or ""
    if nombre:
        normalized["nombre"] = str(nombre).strip()
    
    # Precio (>= 0)
    precio = _normalize_decimal(row.get("precio") or row.get("price"))
    if precio is not None and precio >= 0:
        normalized["precio"] = precio
    
    # Stock (>= 0)
    stock = _normalize_decimal(row.get("stock"))
    if stock is not None and stock >= 0:
        normalized["stock"] = stock
    
    # Proveedor_ID (opcional)
    proveedor_id = row.get("proveedor_id") or row.get("proveedor") or row.get("supplier_id")
    if proveedor_id:
        prov_id = _normalize_int(proveedor_id)
        if prov_id:
            normalized["proveedor_id"] = prov_id
    
    return normalized

