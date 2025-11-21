# app/services/libro_iva_compras_service.py
"""
Servicio para gestión del Libro IVA Compras
ABM manual de facturas de compra + Export CSV/XLSX
"""

from typing import Optional, List, Dict, Any
from datetime import datetime, date
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_, desc
from fastapi import HTTPException
from io import BytesIO
import csv

from app.models.purchase_invoice_model import PurchaseInvoice


class LibroIVAComprasService:
    """Servicio para gestión del Libro IVA Compras"""
    
    def __init__(self, db: Session):
        self.db = db
    
    def get_registros(self, desde: Optional[str] = None, hasta: Optional[str] = None) -> List[PurchaseInvoice]:
        """Obtener registros de IVA Compras con filtros"""
        query = self.db.query(PurchaseInvoice)
        
        if desde:
            query = query.filter(PurchaseInvoice.fecha >= desde)
        if hasta:
            query = query.filter(PurchaseInvoice.fecha <= hasta)
        
        return query.order_by(desc(PurchaseInvoice.fecha)).all()
    
    def export_csv(self, registros: List[PurchaseInvoice]) -> str:
        """Exportar registros a CSV"""
        import io
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            "Fecha", "Proveedor", "Tipo", "PtoVta", "Numero",
            "DocTipo", "DocNro", "Neto", "IVA", "Exento", "Total", "Alicuota"
        ])
        
        for r in registros:
            writer.writerow([
                r.fecha.strftime("%Y-%m-%d") if r.fecha else "",
                r.proveedor_nombre or "",
                r.tipo_cbte,
                r.pto_vta,
                r.nro_cbte,
                r.doc_tipo or "",
                r.doc_nro or "",
                f"{r.imp_neto:.2f}",
                f"{r.imp_iva:.2f}",
                f"{r.imp_exento:.2f}",
                f"{r.imp_total:.2f}",
                f"{r.alicuota_principal:.2f}" if r.alicuota_principal else "0.00"
            ])
        
        return output.getvalue()
    
    def export_xlsx(self, registros: List[PurchaseInvoice]) -> bytes:
        """Exportar registros a XLSX"""
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(status_code=400, detail="openpyxl no instalado")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "IVA Compras"
        
        # Header
        ws.append([
            "Fecha", "Proveedor", "Tipo", "PtoVta", "Numero",
            "DocTipo", "DocNro", "Neto", "IVA", "Exento", "Total", "Alicuota"
        ])
        
        for r in registros:
            ws.append([
                r.fecha.strftime("%Y-%m-%d") if r.fecha else "",
                r.proveedor_nombre or "",
                r.tipo_cbte,
                r.pto_vta,
                r.nro_cbte,
                r.doc_tipo or "",
                r.doc_nro or "",
                float(r.imp_neto),
                float(r.imp_iva),
                float(r.imp_exento),
                float(r.imp_total),
                float(r.alicuota_principal) if r.alicuota_principal else 0.0
            ])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


def crear_purchase_invoice(
    db: Session,
    proveedor_id: Optional[int],
    proveedor_nombre: Optional[str],
    fecha: date,
    tipo_cbte: int,
    pto_vta: int,
    nro_cbte: int,
    doc_tipo: Optional[int],
    doc_nro: Optional[str],
    imp_neto: float,
    imp_iva: float,
    imp_exento: float,
    imp_total: float,
    alicuota_principal: float = 21.0,
    compra_id: Optional[int] = None,
) -> PurchaseInvoice:
    """Crea un registro de factura de compra para el Libro IVA"""
    invoice = PurchaseInvoice(
        proveedor_id=proveedor_id,
        proveedor_nombre=proveedor_nombre,
        fecha=fecha,
        tipo_cbte=tipo_cbte,
        pto_vta=pto_vta,
        nro_cbte=nro_cbte,
        doc_tipo=doc_tipo,
        doc_nro=doc_nro,
        imp_neto=imp_neto,
        imp_iva=imp_iva,
        imp_exento=imp_exento,
        imp_total=imp_total,
        alicuota_principal=alicuota_principal,
        compra_id=compra_id,
    )
    db.add(invoice)
    db.commit()
    db.refresh(invoice)
    return invoice


def actualizar_purchase_invoice(
    db: Session,
    invoice_id: int,
    data: Dict[str, Any],
) -> PurchaseInvoice:
    """Actualiza un registro de factura de compra"""
    invoice = db.query(PurchaseInvoice).filter(PurchaseInvoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura de compra no encontrada")
    
    # Actualizar campos
    for key, value in data.items():
        if hasattr(invoice, key) and value is not None:
            setattr(invoice, key, value)
    
    db.commit()
    db.refresh(invoice)
    return invoice


def eliminar_purchase_invoice(db: Session, invoice_id: int) -> bool:
    """Elimina un registro de factura de compra"""
    invoice = db.query(PurchaseInvoice).filter(PurchaseInvoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Factura de compra no encontrada")
    
    db.delete(invoice)
    db.commit()
    return True


def listar_purchase_invoices(
    db: Session,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    proveedor_id: Optional[int] = None,
    tipo_cbte: Optional[int] = None,
    page: int = 1,
    size: int = 20,
) -> tuple:
    """Lista facturas de compra con filtros"""
    query = db.query(PurchaseInvoice).options(joinedload(PurchaseInvoice.proveedor))
    
    filters = []
    if fecha_desde:
        filters.append(PurchaseInvoice.fecha >= fecha_desde)
    if fecha_hasta:
        filters.append(PurchaseInvoice.fecha <= fecha_hasta)
    if proveedor_id:
        filters.append(PurchaseInvoice.proveedor_id == proveedor_id)
    if tipo_cbte:
        filters.append(PurchaseInvoice.tipo_cbte == tipo_cbte)
    
    if filters:
        query = query.filter(and_(*filters))
    
    total = query.count()
    
    # Ordenar por fecha desc
    query = query.order_by(desc(PurchaseInvoice.fecha))
    
    # Paginación
    offset = (page - 1) * size
    items = query.offset(offset).limit(size).all()
    
    return items, total


def export_libro_iva_compras(
    db: Session,
    fecha_desde: str,
    fecha_hasta: str,
    formato: str = "csv",
    proveedor_id: Optional[int] = None,
    tipo_cbte: Optional[int] = None,
) -> bytes:
    """
    Exporta el Libro IVA Compras en formato CSV o XLSX.
    
    Args:
        db: Sesión de base de datos
        fecha_desde: Fecha desde (YYYY-MM-DD)
        fecha_hasta: Fecha hasta (YYYY-MM-DD)
        formato: Formato de salida ("csv" o "xlsx")
        proveedor_id: Filtro opcional por proveedor
        tipo_cbte: Filtro opcional por tipo de comprobante
    
    Returns:
        Bytes del archivo generado
    """
    # Parsear fechas
    try:
        dt_desde = datetime.strptime(fecha_desde, "%Y-%m-%d").date()
        dt_hasta = datetime.strptime(fecha_hasta, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError("Formato de fecha inválido (YYYY-MM-DD)")
    
    # Consultar facturas de compra
    invoices, _ = listar_purchase_invoices(
        db,
        fecha_desde=dt_desde,
        fecha_hasta=dt_hasta,
        proveedor_id=proveedor_id,
        tipo_cbte=tipo_cbte,
        page=1,
        size=10000,  # Sin límite para export
    )
    
    # Generar datos para el reporte
    rows = []
    for inv in invoices:
        proveedor_nombre = inv.proveedor.nombre if inv.proveedor else (inv.proveedor_nombre or "—")
        doc = inv.doc_nro if inv.doc_nro else "—"
        
        row = {
            "Fecha": inv.fecha.strftime("%d/%m/%Y"),
            "Tipo": _get_tipo_nombre(inv.tipo_cbte),
            "Pto. Vta.": f"{inv.pto_vta:04d}",
            "Nro. Cbte.": f"{inv.nro_cbte:08d}",
            "Proveedor": proveedor_nombre,
            "Doc. Nro.": doc,
            "Neto Gravado": f"{float(inv.imp_neto):.2f}",
            "Exento": f"{float(inv.imp_exento):.2f}",
            "IVA": f"{float(inv.imp_iva):.2f}",
            "Total": f"{float(inv.imp_total):.2f}",
            "Alíc. Principal": f"{float(inv.alicuota_principal):.1f}%",
        }
        rows.append(row)
    
    # Generar archivo según formato
    if formato == "xlsx":
        return _generate_xlsx(rows)
    else:  # csv
        return _generate_csv(rows)


def _get_tipo_nombre(tipo_cbte: int) -> str:
    """Convierte tipo de comprobante numérico a letra"""
    mapping = {
        1: "FA-A",
        6: "FA-B",
        11: "FA-C",
        3: "NC-A",
        8: "NC-B",
        13: "NC-C",
        2: "ND-A",
        7: "ND-B",
        12: "ND-C",
    }
    return mapping.get(tipo_cbte, f"T{tipo_cbte}")


def _generate_csv(rows: List[Dict[str, Any]]) -> bytes:
    """Genera archivo CSV"""
    if not rows:
        return b""
    
    buffer = BytesIO()
    fieldnames = list(rows[0].keys())
    
    from io import StringIO
    string_buffer = StringIO()
    writer = csv.DictWriter(string_buffer, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    
    csv_content = string_buffer.getvalue()
    buffer.write(csv_content.encode('utf-8-sig'))
    buffer.seek(0)
    
    return buffer.getvalue()


def _generate_xlsx(rows: List[Dict[str, Any]]) -> bytes:
    """Genera archivo XLSX"""
    if not rows:
        return b""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise ImportError("openpyxl no está instalado")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro IVA Compras"
    
    # Encabezado
    headers = list(rows[0].keys())
    ws.append(headers)
    
    # Estilo al encabezado
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    for row in rows:
        ws.append(list(row.values()))
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()

