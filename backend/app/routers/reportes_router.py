# app/routers/reportes_router.py
from __future__ import annotations

from datetime import date, datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy import func, extract, case
from sqlalchemy.orm import Session
from io import BytesIO
import csv

from app.core.deps import get_current_user, require_user
from app.db.database import get_db
from app.models.venta_model import Venta, VentaItem
from app.models.compra_model import Compra, CompraItem
from app.models.cliente_model import Cliente
from app.models.proveedor_model import Proveedor
from app.models.producto_model import Producto
from app.models.user_model import User
from app.services.reportes_pedidos_service import reporte_pedidos as get_reporte_pedidos

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

router = APIRouter(prefix="/reportes", tags=["Reportes"])


def _normalize_date(d: Optional[date | datetime | str]) -> Optional[datetime]:
    """Normaliza fecha a datetime con TZ -03:00 (Argentina)"""
    if not d:
        return None
    if isinstance(d, str):
        try:
            d = datetime.fromisoformat(d.replace("Z", "+00:00"))
        except:
            d = datetime.strptime(d, "%Y-%m-%d")
    if isinstance(d, date) and not isinstance(d, datetime):
        d = datetime.combine(d, datetime.min.time())
    # Ajustar a TZ -03:00 si no tiene TZ
    if d.tzinfo is None:
        from datetime import timezone, timedelta
        d = d.replace(tzinfo=timezone(timedelta(hours=-3)))
    return d


@router.get("/ventas")
def reporte_ventas(
    desde: Optional[date] = Query(None, description="Fecha desde"),
    hasta: Optional[date] = Query(None, description="Fecha hasta"),
    group_by: str = Query("dia", description="Agrupar por: dia, cliente, producto"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    """
    Reporte de ventas agrupado.
    Respuesta: {items: [{clave, cantidad_items, total_cantidad, total_monto}], total_general: {...}}
    """
    desde_dt = _normalize_date(desde)
    hasta_dt = _normalize_date(hasta)
    
    # Query base
    query = db.query(Venta).join(VentaItem, VentaItem.venta_id == Venta.id)
    
    # Filtros de fecha
    if desde_dt:
        query = query.filter(Venta.fecha >= desde_dt)
    if hasta_dt:
        # Incluir todo el día hasta
        hasta_end = datetime.combine(hasta, datetime.max.time())
        if hasta_dt.tzinfo:
            hasta_end = hasta_end.replace(tzinfo=hasta_dt.tzinfo)
        query = query.filter(Venta.fecha <= hasta_end)
    
    # Agrupaciones
    group_cols = []
    if group_by == "dia":
        # Agrupar por día (sin hora)
        group_cols.append(func.date(Venta.fecha).label("clave"))
    elif group_by == "cliente":
        query = query.outerjoin(Cliente, Cliente.id == Venta.cliente_id)
        group_cols.append(func.coalesce(Cliente.nombre, "Sin cliente").label("clave"))
    elif group_by == "producto":
        query = query.outerjoin(Producto, Producto.id == VentaItem.producto_id)
        group_cols.append(func.coalesce(Producto.nombre, "Sin producto").label("clave"))
    else:
        # Default: día
        group_cols.append(func.date(Venta.fecha).label("clave"))
    
    # Agregaciones
    cantidad_items = func.count(Venta.id.distinct()).label("cantidad_items")
    total_cantidad = func.sum(VentaItem.cantidad).label("total_cantidad")
    total_monto = func.sum(VentaItem.subtotal).label("total_monto")
    
    result = query.group_by(*group_cols).with_entities(
        *group_cols,
        cantidad_items,
        total_cantidad,
        total_monto,
    ).all()
    
    items = []
    total_general = {"cantidad_items": 0, "total_cantidad": 0.0, "total_monto": 0.0}
    
    for row in result:
        clave = str(row.clave) if row.clave else "Sin agrupar"
        items.append({
            "clave": clave,
            "cantidad_items": row.cantidad_items or 0,
            "total_cantidad": float(row.total_cantidad or 0),
            "total_monto": float(row.total_monto or 0),
        })
        total_general["cantidad_items"] += row.cantidad_items or 0
        total_general["total_cantidad"] += float(row.total_cantidad or 0)
        total_general["total_monto"] += float(row.total_monto or 0)
    
    return {
        "items": items,
        "total_general": total_general,
        "group_by": group_by,
        "desde": desde.isoformat() if desde else None,
        "hasta": hasta.isoformat() if hasta else None,
    }


@router.get("/compras")
def reporte_compras(
    desde: Optional[date] = Query(None, description="Fecha desde"),
    hasta: Optional[date] = Query(None, description="Fecha hasta"),
    group_by: str = Query("dia", description="Agrupar por: dia, proveedor, producto"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    """
    Reporte de compras agrupado.
    Respuesta: {items: [{clave, cantidad_items, total_cantidad, total_monto}], total_general: {...}}
    """
    desde_dt = _normalize_date(desde)
    hasta_dt = _normalize_date(hasta)
    
    # Query base
    query = db.query(Compra).join(CompraItem, CompraItem.compra_id == Compra.id)
    
    # Filtros de fecha
    if desde_dt:
        query = query.filter(Compra.fecha >= desde_dt)
    if hasta_dt:
        hasta_end = datetime.combine(hasta, datetime.max.time())
        if hasta_dt.tzinfo:
            hasta_end = hasta_end.replace(tzinfo=hasta_dt.tzinfo)
        query = query.filter(Compra.fecha <= hasta_end)
    
    # Agrupaciones
    group_cols = []
    if group_by == "dia":
        group_cols.append(func.date(Compra.fecha).label("clave"))
    elif group_by == "proveedor":
        query = query.outerjoin(Proveedor, Proveedor.id == Compra.proveedor_id)
        group_cols.append(func.coalesce(Proveedor.nombre, "Sin proveedor").label("clave"))
    elif group_by == "producto":
        query = query.outerjoin(Producto, Producto.id == CompraItem.producto_id)
        group_cols.append(func.coalesce(Producto.nombre, "Sin producto").label("clave"))
    else:
        group_cols.append(func.date(Compra.fecha).label("clave"))
    
    # Agregaciones
    cantidad_items = func.count(Compra.id.distinct()).label("cantidad_items")
    total_cantidad = func.sum(CompraItem.cantidad).label("total_cantidad")
    total_monto = func.sum(CompraItem.subtotal).label("total_monto")
    
    result = query.group_by(*group_cols).with_entities(
        *group_cols,
        cantidad_items,
        total_cantidad,
        total_monto,
    ).all()
    
    items = []
    total_general = {"cantidad_items": 0, "total_cantidad": 0.0, "total_monto": 0.0}
    
    for row in result:
        clave = str(row.clave) if row.clave else "Sin agrupar"
        items.append({
            "clave": clave,
            "cantidad_items": row.cantidad_items or 0,
            "total_cantidad": float(row.total_cantidad or 0),
            "total_monto": float(row.total_monto or 0),
        })
        total_general["cantidad_items"] += row.cantidad_items or 0
        total_general["total_cantidad"] += float(row.total_cantidad or 0)
        total_general["total_monto"] += float(row.total_monto or 0)
    
    return {
        "items": items,
        "total_general": total_general,
        "group_by": group_by,
        "desde": desde.isoformat() if desde else None,
        "hasta": hasta.isoformat() if hasta else None,
    }


@router.get("/ventas/export")
def exportar_ventas(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    group_by: str = Query("dia"),
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    """Exportar reporte de ventas a CSV o XLSX"""
    data = reporte_ventas(desde=desde, hasta=hasta, group_by=group_by, db=db, current_user=current_user)
    
    if format == "csv":
        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow(["Clave", "Cantidad de Ventas", "Total Cantidad", "Total Monto"])
        for item in data["items"]:
            writer.writerow([
                item["clave"],
                item["cantidad_items"],
                item["total_cantidad"],
                item["total_monto"],
            ])
        writer.writerow([
            "TOTAL",
            data["total_general"]["cantidad_items"],
            data["total_general"]["total_cantidad"],
            data["total_general"]["total_monto"],
        ])
        output.seek(0)
        return Response(
            content=output.read(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="reporte_ventas_{desde or "all"}_{hasta or "all"}.csv"'}
        )
    elif format == "xlsx" and Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"
        ws.append(["Clave", "Cantidad de Ventas", "Total Cantidad", "Total Monto"])
        for item in data["items"]:
            ws.append([
                item["clave"],
                item["cantidad_items"],
                item["total_cantidad"],
                item["total_monto"],
            ])
        ws.append([
            "TOTAL",
            data["total_general"]["cantidad_items"],
            data["total_general"]["total_cantidad"],
            data["total_general"]["total_monto"],
        ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="reporte_ventas_{desde or "all"}_{hasta or "all"}.xlsx"'}
        )
    else:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=400, detail="XLSX requiere openpyxl instalado")


@router.get("/compras/export")
def exportar_compras(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    group_by: str = Query("dia"),
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    """Exportar reporte de compras a CSV o XLSX"""
    data = reporte_compras(desde=desde, hasta=hasta, group_by=group_by, db=db, current_user=current_user)
    
    if format == "csv":
        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow(["Clave", "Cantidad de Compras", "Total Cantidad", "Total Monto"])
        for item in data["items"]:
            writer.writerow([
                item["clave"],
                item["cantidad_items"],
                item["total_cantidad"],
                item["total_monto"],
            ])
        writer.writerow([
            "TOTAL",
            data["total_general"]["cantidad_items"],
            data["total_general"]["total_cantidad"],
            data["total_general"]["total_monto"],
        ])
        output.seek(0)
        return Response(
            content=output.read(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="reporte_compras_{desde or "all"}_{hasta or "all"}.csv"'}
        )
    elif format == "xlsx" and Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Compras"
        ws.append(["Clave", "Cantidad de Compras", "Total Cantidad", "Total Monto"])
        for item in data["items"]:
            ws.append([
                item["clave"],
                item["cantidad_items"],
                item["total_cantidad"],
                item["total_monto"],
            ])
        ws.append([
            "TOTAL",
            data["total_general"]["cantidad_items"],
            data["total_general"]["total_cantidad"],
            data["total_general"]["total_monto"],
        ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="reporte_compras_{desde or "all"}_{hasta or "all"}.xlsx"'}
        )
    else:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=400, detail="XLSX requiere openpyxl instalado")


@router.get("/pedidos")
def reporte_pedidos_endpoint(
    desde: Optional[date] = Query(None, description="Fecha desde"),
    hasta: Optional[date] = Query(None, description="Fecha hasta"),
    group_by: str = Query("estado", description="Agrupar por: estado, dia, cliente"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    """
    Reporte de pedidos agrupado por estado, día o cliente
    """
    desde_dt = _normalize_date(desde) if desde else None
    hasta_dt = _normalize_date(hasta) if hasta else None
    
    return get_reporte_pedidos(db, desde=desde_dt, hasta=hasta_dt, group_by=group_by)


@router.get("/pedidos/export")
def exportar_pedidos(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    group_by: str = Query("estado"),
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    """Exportar reporte de pedidos a CSV o XLSX"""
    desde_dt = _normalize_date(desde) if desde else None
    hasta_dt = _normalize_date(hasta) if hasta else None
    data = get_reporte_pedidos(db, desde=desde_dt, hasta=hasta_dt, group_by=group_by)
    
    if format == "csv":
        output = BytesIO()
        writer = csv.writer(output)
        writer.writerow(["Grupo", "Cantidad", "Total"])
        for item in data["items"]:
            writer.writerow([
                item["grupo"],
                item["cantidad"],
                item["total"],
            ])
        output.seek(0)
        return Response(
            content=output.read(),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="reporte_pedidos_{group_by}_{desde or "all"}_{hasta or "all"}.csv"'}
        )
    elif format == "xlsx" and Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pedidos"
        ws.append(["Grupo", "Cantidad", "Total"])
        for item in data["items"]:
            ws.append([
                item["grupo"],
                item["cantidad"],
                item["total"],
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            content=output.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="reporte_pedidos_{group_by}_{desde or "all"}_{hasta or "all"}.xlsx"'}
        )
    else:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=400, detail="XLSX requiere openpyxl instalado")


@router.get("/libro-iva-ventas")
def libro_iva_ventas(
    desde: str = Query(..., description="Fecha desde (YYYY-MM-DD)"),
    hasta: str = Query(..., description="Fecha hasta (YYYY-MM-DD)"),
    format: str = Query("csv", regex="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
):
    """
    Genera el Libro IVA Ventas en formato CSV o XLSX.
    
    Incluye todas las facturas electrónicas emitidas en el período especificado,
    con los datos requeridos para la presentación de IVA ante AFIP.
    
    - **desde**: Fecha desde (YYYY-MM-DD)
    - **hasta**: Fecha hasta (YYYY-MM-DD)
    - **format**: Formato de salida (csv o xlsx)
    """
    from app.services.libro_iva_ventas_service import generar_libro_iva_ventas
    
    try:
        archivo_bytes = generar_libro_iva_ventas(db, desde, hasta, format)
        
        media_type = "text/csv" if format == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = format
        
        return Response(
            content=archivo_bytes,
            media_type=media_type,
            headers={"Content-Disposition": f'attachment; filename="libro_iva_ventas_{desde}_{hasta}.{extension}"'}
        )
    except ValueError as e:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=400, detail=str(e))

